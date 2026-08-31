"""
Tests for the standalone runner script.

The script exists so the model can be run without a notebook, and the whole
point of it is that it does *the same thing* the notebook does. So the load
bearing tests here are not "it produces a file" -- they are:

* every stage actually ran (a script that silently skips the optimizer and
  still writes a workbook is worse than one that crashes),
* the workbook it writes has the same nine sheets as the notebook's,
* the proposals JSON lands under ``propuestas/`` and carries no internal keys,
* and the command-line flags reach the model rather than being parsed and
  ignored -- which is the failure mode that leaves someone running Moderado
  for a week while believing they set Agresivo.

The Yahoo download is the one thing that cannot run here, so it is stubbed
exactly as ``test_notebook.py`` stubs it, against the same fixture.
"""

from __future__ import annotations

import io
import json
import sys
import tempfile
from contextlib import redirect_stdout
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import correr_modelo  # noqa: E402
import screener.yahoo_adapter as _ya  # noqa: E402
from screener.black_litterman import DRIVE_PROPOSALS_DIR  # noqa: E402
from screener.tuning import reset_all  # noqa: E402
from screener.yahoo_adapter import build_market_data  # noqa: E402
from test_yahoo_adapter import make_yf_frame  # noqa: E402

PASSED = 0
FAILED = 0

TICKERS = ["SPY", "QQQ", "IWM", "GLD", "DBC", "TLT", "LQD", "BIL", "HYG",
           "XLE", "XLV", "XLF", "AAPL", "MSFT", "NVDA", "JPM", "LLY"]


def check(label: str, condition: bool, detail: str = "") -> None:
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"PASS {label}")
    else:
        FAILED += 1
        print(f"FAIL {label}" + (f"  -- {detail}" if detail else ""))


# --------------------------------------------------------------------------
# Offline harness
# --------------------------------------------------------------------------

def run_script(*argv: str) -> tuple[str, Path]:
    """
    Run the script's main() with the network calls stubbed.

    Returns its stdout and the output directory. The stubs replace only the two
    functions that reach Yahoo; everything downstream is the real code path.
    """
    tmp = Path(tempfile.mkdtemp(prefix="correr-"))

    real_fetch = _ya.fetch_market_data
    real_caps = _ya.fetch_market_caps
    real_universe = _ya.default_universe

    def fake_fetch(tickers, *, benchmark="SPY", risk_free_rate=0.0425, **kw):
        frame = make_yf_frame(list(tickers), dividends={"SPY": 6.0, "JPM": 4.0})
        # Thin volume on one name so the per-profile liquidity floors actually
        # differ, which is what makes the comparison table exercise `n/e`.
        if ("Volume", "DBC") in frame.columns:
            frame[("Volume", "DBC")] = 250_000.0
        data = build_market_data(frame, list(tickers), benchmark=benchmark,
                                 risk_free_rate=risk_free_rate)
        return (data, frame) if kw.get("with_frame") else data

    def fake_caps(tickers, **kw):
        return {t: 1e10 + 1e9 * i for i, t in enumerate(tickers)}

    _ya.fetch_market_data = fake_fetch
    _ya.fetch_market_caps = fake_caps
    _ya.default_universe = lambda groups, benchmark="SPY": list(TICKERS)
    # The script imports these names into its own frame at call time, so the
    # module-level patch above is what it will actually see.
    buffer = io.StringIO()
    try:
        with redirect_stdout(buffer):
            code = correr_modelo.main([*argv, "--salida", str(tmp)])
    finally:
        _ya.fetch_market_data = real_fetch
        _ya.fetch_market_caps = real_caps
        _ya.default_universe = real_universe
        reset_all()

    assert code == 0, f"script returned {code}"
    return buffer.getvalue(), tmp


# --------------------------------------------------------------------------
# Tests
# --------------------------------------------------------------------------

def test_every_stage_runs() -> None:
    """
    A run that quietly skips a stage and still writes a workbook is the
    failure this guards against.
    """
    out, tmp = run_script()

    for stage in ("1 · UNIVERSO", "2 · DESCARGA", "3 · COBERTURA",
                  "4 · SCREENING", "5 · COMPARACIÓN", "6 · VIEWS",
                  "7 · DIAGNÓSTICOS", "8 · CARTERA", "9 · ARCHIVOS"):
        check(f"stage ran: {stage}", stage in out,
              "missing from stdout")

    check("the optimizer reported a status",
          "estado: optimal" in out, out[-600:])
    check("the band audit ran and is reported",
          "Auditoría de bandas" in out or "AUDITORÍA" in out)
    check("both diagnostics printed",
          "Correlación entre bloques" in out and "Saturación de views" in out)
    check("the run reports its equilibrium anchor by asset class",
          "Ancla (politica) por clase de activo" in out)


def test_workbook_matches_the_notebook_shape() -> None:
    """Same nine sheets the notebook writes, in the same order."""
    from openpyxl import load_workbook

    _, tmp = run_script()
    book = tmp / "screening.xlsx"
    check("the workbook was written", book.exists())

    wb = load_workbook(book)
    expected = ["Ranking", "Bloques", "Perfiles", "Views BL",
                "Cartera", "Cesta", "Universo", "Cobertura", "Parametros"]
    check("nine sheets, matching the notebook", wb.sheetnames == expected,
          str(wb.sheetnames))
    check("the Cartera sheet is never blank -- positions or a stated reason",
          wb["Cartera"].max_row > 1)
    check("the Ranking sheet has one row per scored name",
          wb["Ranking"].max_row > 5, f"{wb['Ranking'].max_row} rows")

    params = {row[0].value: row[1].value
              for row in wb["Parametros"].iter_rows(min_row=2)}
    check("Parametros records that no account was read",
          params.get("Portafolio", "").startswith("ninguno"))
    check("Parametros records the anchor actually used",
          params.get("Ancla del equilibrio") == "politica",
          str(params.get("Ancla del equilibrio")))
    check("Parametros records the absolute quality floor",
          "Sharpe" in str(params.get("Piso absoluto para Overweight", "")),
          str(params.get("Piso absoluto para Overweight")))
    check("Parametros still flags the IC as an assumption",
          "no calibrado" in str(params.get("Nota sobre el IC", "")))
    check("Parametros says midpoints are not a real SAA",
          "Comité de Inversiones" in str(params.get("Nota sobre el ancla", "")),
          str(params.get("Nota sobre el ancla"))[:120])


def test_proposals_land_in_the_right_folder() -> None:
    """
    The governance boundary. Proposals go to propuestas/, never aprobadas/,
    and the internal diagnostic key must not reach the file a manager signs.
    """
    _, tmp = run_script()
    proposals = list((tmp / DRIVE_PROPOSALS_DIR).glob("*.json"))
    check("the proposals JSON was written under propuestas/",
          len(proposals) == 1, str(list(tmp.rglob('*.json'))))

    payload = json.loads(proposals[0].read_text(encoding="utf-8"))
    check("the file names the strategy it targets",
          payload.get("estrategia") == "Moderado")
    check("the file states the screen read no account",
          "sin datos de cuenta" in payload.get("origen", ""))
    check("no internal diagnostic keys reached the file",
          all(not k.startswith("_") for v in payload["views"] for k in v))
    check("the calibration block travels with the views",
          "calibracion" in payload and
          "nota" in payload["calibracion"])

    _, tmp2 = run_script("--sin-json")
    check("--sin-json suppresses the proposals file",
          not list(tmp2.rglob("*.json")),
          str(list(tmp2.rglob("*.json"))))


def test_flags_actually_reach_the_model() -> None:
    """
    A flag that parses but never reaches the model is the worst kind of bug
    here: the run looks right and is scored under a different mandate.
    """
    from openpyxl import load_workbook

    _, tmp = run_script("--perfil", "Agresivo", "--estrategia", "Agresivo",
                        "--ic", "0.03", "--top-n", "12", "--max-views", "3")
    params = {row[0].value: row[1].value
              for row in load_workbook(tmp / "screening.xlsx")["Parametros"]
              .iter_rows(min_row=2)}

    check("--perfil reached the screen", params.get("Perfil") == "Agresivo",
          str(params.get("Perfil")))
    check("--estrategia reached the export",
          params.get("Estrategia CCI destino") == "Agresivo")
    check("--ic reached the view translation", params.get("IC supuesto (views)") == 0.03,
          str(params.get("IC supuesto (views)")))
    check("--perfil rewired the gates, not just the label",
          params.get("Beta máxima") == "1.80", str(params.get("Beta máxima")))
    check("--perfil rewired the block weights",
          params.get("Peso — Momentum & Trend") == "36%",
          str(params.get("Peso — Momentum & Trend")))

    proposals = json.loads(
        next((tmp / DRIVE_PROPOSALS_DIR).glob("*.json")).read_text(encoding="utf-8"))
    check("--max-views capped the export", len(proposals["views"]) <= 3,
          f"{len(proposals['views'])} views")


def test_market_anchor_flag_changes_the_anchor() -> None:
    """Both anchors must run, so the two can actually be compared."""
    out, tmp = run_script("--ancla", "mercado")
    check("--ancla mercado runs end to end", "estado: optimal" in out)
    check("the market anchor is reported as such",
          "Ancla (mercado) por clase de activo" in out)

    from openpyxl import load_workbook
    params = {row[0].value: row[1].value
              for row in load_workbook(tmp / "screening.xlsx")["Parametros"]
              .iter_rows(min_row=2)}
    check("the workbook records which anchor was used",
          params.get("Ancla del equilibrio") == "mercado")


def test_custom_ticker_list() -> None:
    """--tickers implies the custom universe and warns about the product filter."""
    args = correr_modelo.parse_args(["--tickers", "SPY,QQQ,TLT"])
    check("--tickers switches the universe to the custom list",
          args.universo == "lista", args.universo)

    out, _ = run_script("--tickers", "SPY,QQQ,IWM,GLD,TLT,LQD,BIL,AAPL,MSFT,JPM")
    check("a custom list runs end to end", "9 · ARCHIVOS" in out)
    check("it warns that the leveraged-product filter is blind without names",
          "AVISO" in out and "apalancado" in out)


def test_profile_and_strategy_mismatch_is_flagged() -> None:
    """
    Screening under one mandate and exporting for another is legal but almost
    always a mistake, so it has to be said out loud.
    """
    out, _ = run_script("--perfil", "Conservador", "--estrategia", "Agresivo")
    check("a profile/strategy mismatch is called out",
          "AVISO" in out and "otro mandato" in out,
          out[:400])


def main() -> int:
    for fn in [
        test_every_stage_runs,
        test_workbook_matches_the_notebook_shape,
        test_proposals_land_in_the_right_folder,
        test_flags_actually_reach_the_model,
        test_market_anchor_flag_changes_the_anchor,
        test_custom_ticker_list,
        test_profile_and_strategy_mismatch_is_flagged,
    ]:
        fn()

    print("-" * 70)
    print(f"{PASSED}/{PASSED + FAILED} passed")
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
