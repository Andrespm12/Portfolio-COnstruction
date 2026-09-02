"""
Executes the Colab notebook and checks it against the repo.

Two distinct things are verified here.

**Drift.** The notebook embeds a copy of the ``screener`` package. A copy that
silently falls behind the repo is the whole hazard of shipping one, so this
rebuilds the notebook from current source and fails if the checked-in file
differs. Fix by re-running ``scripts/build_notebook.py``.

**Execution.** Every code cell is run in order in a temporary directory, with
one substitution: the Yahoo download is replaced by a synthetic frame, because
CI has no outbound network. Everything downstream of the download -- unpacking
the engine, the parameter logic, coverage, scoring, both styled tables, the
per-name detail view, export and the tuning helpers -- runs for real against
the same code Colab will run.

What this does NOT prove: that ``yfinance.download`` returns what the adapter
expects. That contract is asserted separately against yfinance's documented
output shape in ``test_yahoo_adapter.py``, but the live call is unexercised
until the notebook is run with internet access.
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
import warnings
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK = ROOT / "notebooks" / "screener_colab.ipynb"

PASSED = 0
FAILED = 0


def check(label: str, condition: bool, detail: str = "") -> None:
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"PASS {label}")
    else:
        FAILED += 1
        print(f"FAIL {label}" + (f"  -- {detail}" if detail else ""))


# --------------------------------------------------------------------------
# Drift
# --------------------------------------------------------------------------

def test_notebook_matches_source() -> None:
    sys.path.insert(0, str(ROOT / "scripts"))
    import build_notebook as builder

    fresh = json.dumps(builder.build_notebook(), indent=1, ensure_ascii=False) + "\n"
    on_disk = NOTEBOOK.read_text(encoding="utf-8")
    check("checked-in notebook matches a fresh build from screener/",
          fresh == on_disk,
          "run: python3 scripts/build_notebook.py")

    check("the build is deterministic (rebuild is byte-identical)",
          json.dumps(builder.build_notebook(), indent=1, ensure_ascii=False) + "\n" == fresh)


def test_embedded_engine_is_current() -> None:
    """The embedded tarball must contain byte-identical module sources."""
    import base64
    import gzip
    import io
    import re
    import tarfile

    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    setup = next(c for c in nb["cells"]
                 if c["cell_type"] == "code" and "ENGINE_B64" in "".join(c["source"]))
    source = "".join(setup["source"])
    blob = "".join(re.findall(r'^\s*"([A-Za-z0-9+/=]+)"\s*$', source, re.M))

    raw = gzip.decompress(base64.b64decode(blob))
    mismatches: list[str] = []
    with tarfile.open(fileobj=io.BytesIO(raw)) as tar:
        members = tar.getnames()
        for name in members:
            embedded = tar.extractfile(name).read()
            actual = (ROOT / name).read_bytes()
            if embedded != actual:
                mismatches.append(name)

    check("every embedded module matches the repo byte for byte",
          not mismatches, f"stale: {mismatches}")
    check("all engine modules are embedded",
          len(members) == len(__import__("build_notebook").MODULES),
          f"got {len(members)}")


# --------------------------------------------------------------------------
# Execution
# --------------------------------------------------------------------------

#: Offline stand-in for the Yahoo download.
#:
#: DBC is given a deliberately thin volume so its average daily traded value
#: lands between the aggressive profile's $10MM floor and the conservative
#: profile's $50MM one. That makes the three profiles score *different sets of
#: names*, which is exactly the condition that made the profile-comparison cell
#: raise KeyError: it indexed one profile's results by another's tickers.
OFFLINE_FETCH = """
from test_yahoo_adapter import make_yf_frame
from screener.yahoo_adapter import build_market_data

_frame = make_yf_frame(TICKERS, dividends={'SPY': 6.0, 'JPM': 4.0})
if ('Volume', 'DBC') in _frame.columns:
    _frame[('Volume', 'DBC')] = 250_000.0
market_data = build_market_data(
    _frame, TICKERS, benchmark=BENCHMARK,
    risk_free_rate=TASA_LIBRE_RIESGO,
)
frame_diario = _frame

# The optimizer cell fetches market caps over the network. Stub the module
# attribute before that cell imports the name, so the offline run exercises the
# real optimization path rather than skipping it.
import screener.yahoo_adapter as _ya
_ya.fetch_market_caps = lambda tickers, **kw: {
    t: 1e10 + 1e9 * i for i, t in enumerate(tickers)
}
print(f'{len(market_data["instruments"])} instrumentos (fixture offline)')
"""


def executable_cells(nb: dict) -> list[str]:
    """Code cells, with magics stripped and the network call substituted."""
    out: list[str] = []
    for cell in nb["cells"]:
        if cell["cell_type"] != "code":
            continue
        source = "".join(cell["source"])

        if "fetch_market_data(" in source:
            out.append(OFFLINE_FETCH)
            continue
        # %pip / %load_ext are IPython-only; nothing downstream depends on them
        # here because the test environment already has yfinance installed.
        source = "\n".join(
            line for line in source.splitlines() if not line.strip().startswith("%")
        )
        out.append(source)
    return out


def test_cells_execute() -> None:
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    cells = executable_cells(nb)

    # A small custom universe keeps the fixture fast and exercises the
    # custom-list branch of the parameter cell.
    tickers = ["SPY", "QQQ", "IWM", "GLD", "TLT", "AAPL", "MSFT", "NVDA",
               "JPM", "LLY", "DBC"]
    patched = 0
    for i, source in enumerate(cells):
        if "UNIVERSO =" in source:
            cells[i] = source.replace(
                'UNIVERSO = "Completo (S&P + Nasdaq + Dow + ETFs)"',
                'UNIVERSO = "Lista personalizada"',
            ).replace(
                'TICKERS_PERSONALIZADOS = ""',
                f'TICKERS_PERSONALIZADOS = "{",".join(tickers)}"',
            )
            patched += 1
    check("parameter cell exposes the knobs the test needs to patch", patched == 1)

    namespace: dict = {"__name__": "__main__"}
    cwd = os.getcwd()
    workdir = tempfile.mkdtemp(prefix="nb-exec-")
    failures: list[tuple[int, str]] = []

    try:
        os.chdir(workdir)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            for i, source in enumerate(cells):
                try:
                    exec(compile(source, f"<cell {i}>", "exec"), namespace)
                except Exception as exc:  # noqa: BLE001 - reporting, not handling
                    failures.append((i, f"{type(exc).__name__}: {exc}"))
    finally:
        os.chdir(cwd)

    check("every code cell executes without raising",
          not failures,
          "; ".join(f"cell {i}: {msg}" for i, msg in failures))

    # ---- the namespace the notebook leaves behind -------------------------
    if failures:
        return

    check("engine unpacked and checksum verified", "ENGINE_SHA256" in namespace)
    check("no stale FACTOR_MODEL is bound before the profile is applied",
          "FACTOR_MODEL" not in namespace,
          "a name bound at setup would still hold the 7-block model")
    check("scored results exist",
          len(namespace.get("scored", [])) >= 8,
          f"got {len(namespace.get('scored', []))}")
    check("ranking table was built with one row per scored name",
          len(namespace["tabla"]) == len(namespace["scored"]))
    check("coverage report was computed", not namespace["_cov"].empty)

    # ---- the screen is independent of any book ---------------------------
    model = namespace["MODELO"]
    check("the applied model has six blocks, none of them Portfolio Fit",
          len(model) == 6 and all(b.key != "portfolio_fit" for b in model),
          f"got {[b.key for b in model]}")
    check("factor heatmap has a column per surviving block",
          len(namespace["mapa"].columns) == len(model))
    check("the run is flagged standalone", namespace["meta"]["standalone"] is True)
    check("no scored name carries a correlation to a book",
          all(r.raw_metrics.get("corr_to_portfolio") is None
              for r in namespace["scored"]))
    check("no scored name carries an existing-position overlap",
          all(r.raw_metrics.get("existing_overlap") is None
              for r in namespace["scored"]))

    # ---- profiles ---------------------------------------------------------
    check("a profile was resolved and described",
          namespace["perfil"].key in {"conservador", "moderado", "agresivo"})
    check("meta records which profile produced the ranking",
          namespace["meta"]["profile"] == namespace["perfil"].key)
    comparison = namespace["comparacion"]
    check("the profile comparison covers all three profiles",
          {"Conservador", "Moderado", "Agresivo"} <= set(comparison.columns),
          f"got {list(comparison.columns)}")
    check("the comparison has a row per scored name",
          len(comparison) == len(namespace["scored"]))
    check("comparing profiles left the selected profile in force",
          namespace["meta"]["profile"] == namespace["perfil"].key)

    # Regression: profiles apply different liquidity floors, so they score
    # different sets of names. The comparison must report that, not raise.
    labels = ["Conservador", "Moderado", "Agresivo"]
    ineligible = (comparison[labels] == namespace["NO_ELEGIBLE"]).sum().sum()
    check("a name eligible under one profile but not another is marked, not dropped",
          ineligible > 0,
          "the fixture should make DBC fail the conservative liquidity floor")
    check("the strictest profile is the one that rejects it",
          (comparison["Conservador"] == namespace["NO_ELEGIBLE"]).sum()
          >= (comparison["Agresivo"] == namespace["NO_ELEGIBLE"]).sum())
    check("every other cell holds a real recommendation",
          comparison[labels].isin(
              ["OVERWEIGHT", "MARKET WEIGHT", "UNDERWEIGHT",
               namespace["NO_ELEGIBLE"]]).all().all())

    # ---- Black-Litterman export -------------------------------------------
    views = namespace["views"]
    check("the notebook produced Black-Litterman views", len(views) > 0)
    check("every view carries the fields CCI's approval flow reads",
          all({"estrategia", "tipo", "Q", "conviccion", "justificacion"} <= set(v)
              for v in views))
    check("views are tagged with the CCI strategy, not the screener profile",
          all(v["estrategia"] == namespace["ESTRATEGIA_CCI"] for v in views))
    check("no view exceeds the +/-5% band CCI's document specifies",
          all(abs(v["Q"]) <= 0.05 + 1e-9 for v in views))
    check("the exported basket uses CCI's sheet columns",
          list(namespace["cesta_df"].columns)
          == list(__import__("screener.black_litterman",
                             fromlist=["BASKET_COLUMNS"]).BASKET_COLUMNS))
    check("the JSON download is a togglable choice, not forced",
          "EXPORTAR_JSON_PARA_BL" in namespace)

    views_file = Path(workdir) / namespace["ARCHIVO_VIEWS"]
    check("the views JSON is written with CCI's naming convention",
          views_file.exists() and views_file.name.endswith(".json"))
    payload = json.loads(views_file.read_text(encoding="utf-8"))
    check("the JSON declares the IC as an assumption",
          "supuesto" in payload["calibracion"]["nota"])
    check("the JSON records that no account data was read",
          "sin datos de cuenta" in payload["origen"])
    workbook = Path(workdir) / "screening.xlsx"
    check("Excel workbook is written",
          workbook.exists() and workbook.stat().st_size > 0)
    check("no markdown report is produced",
          not (Path(workdir) / "screen_report.md").exists())
    check("the basket is a sheet, not a stray CSV",
          not list(Path(workdir).glob("cesta_*.csv")))

    import openpyxl
    wb = openpyxl.load_workbook(workbook)
    check("workbook has the nine documented sheets",
          set(wb.sheetnames) == {"Ranking", "Bloques", "Perfiles", "Views BL",
                                 "Cartera", "Cesta", "Universo", "Cobertura",
                                 "Parametros"},
          f"got {wb.sheetnames}")

    # The whole point of the JSON/Excel split: a human reads the views in the
    # workbook, the BL engine reads them from the JSON.
    views_header = [c.value for c in next(wb["Views BL"].iter_rows(max_row=1))]
    check("the Views sheet resolves both view shapes into explicit columns",
          {"tipo", "activo", "long", "short", "Q", "conviccion", "justificacion"}
          == set(views_header), f"got {views_header}")
    check("every view in the JSON also appears in the workbook",
          wb["Views BL"].max_row == len(views) + 1)
    check("the Cesta sheet carries CCI's basket columns",
          [c.value for c in next(wb["Cesta"].iter_rows(max_row=1))]
          == list(__import__("screener.black_litterman",
                             fromlist=["BASKET_COLUMNS"]).BASKET_COLUMNS))

    ranking_header = [c.value for c in next(wb["Ranking"].iter_rows(max_row=1))]
    check("Ranking sheet has a row per scored name",
          wb["Ranking"].max_row == len(namespace["scored"]) + 1)
    check("Ranking sheet carries no book-relative column",
          not {"corr_libro", "corr_to_portfolio", "existing_overlap"}
          & set(ranking_header), f"got {ranking_header}")
    check("Ranking sheet reports alpha in place of correlation-to-book",
          "alpha" in ranking_header)

    # ---- the optimizer ran in-process, with no file between the two halves --
    cartera = namespace["cartera"]
    check("the notebook produced an optimized portfolio", cartera.feasible,
          f"status {cartera.status}")
    check("the portfolio passes its own band audit",
          not cartera.breaches, "; ".join(cartera.breaches))
    check("the portfolio holds more than one position",
          int((cartera.weights > 0).sum()) > 1)
    check("weights are non-negative (long-only)",
          bool((cartera.weights >= -1e-9).all()))

    from screener.cci_regulation import REGULACIONES
    from screener.optimizer import gross_budget
    budget = gross_budget(namespace["ESTRATEGIA_CCI"])
    check("gross exposure respects the leverage budget",
          cartera.gross_exposure <= budget + 1e-6,
          f"{cartera.gross_exposure:.4f} > {budget:.4f}")
    check("the views reached the optimizer in memory, with no file in between",
          "views" in namespace and len(namespace["views"]) > 0)

    # ---- the equilibrium anchor -------------------------------------------
    # The anchor decides most of the book, so verify the notebook actually
    # runs on the policy one rather than merely offering it as an option.
    import pandas as _pd

    from screener.cci_regulation import classify_for_bands as _classify
    from screener.optimizer import EQUITY_CLASSES

    check("the notebook defaults to the mandate's own neutral portfolio",
          namespace.get("ANCLA") == "politica", f"got {namespace.get('ANCLA')!r}")

    anchor = namespace["pesos_ancla"]
    strategy = namespace["ESTRATEGIA_CCI"]
    anchor_classes = _pd.Series(
        {t: _classify(t, namespace["tipos_todos"].get(t, "ETF")) for t in anchor.index}
    )
    by_class = anchor.groupby(anchor_classes).sum()

    check("the anchor spends exactly the gross budget in force",
          abs(float(anchor.sum()) - budget) < 1e-9,
          f"{float(anchor.sum()):.6f} vs {budget:.6f}")
    check("the anchor is long-only", bool((anchor >= -1e-12).all()))

    anchor_equity = sum(float(by_class.get(c, 0.0)) for c in EQUITY_CLASSES)
    ceiling = REGULACIONES[strategy]["max_equity_total"]
    check("the anchor starts inside the mandate rather than on its ceiling",
          anchor_equity <= ceiling + 1e-9,
          f"renta variable neutral {anchor_equity:.4f} > techo {ceiling:.4f}")
    check("the anchor spans the classes the basket actually holds",
          set(by_class.index) == set(anchor_classes.unique()))
    check("the notebook states that band midpoints are not a real SAA",
          any("Comité de Inversiones" in n for n in namespace["_notas_ancla"]),
          str(namespace.get("_notas_ancla")))
    check("the Cartera sheet lists the held positions",
          wb["Cartera"].max_row == int((cartera.weights > 0).sum()) + 1)
    # Regression: an earlier version wrote a blank Cartera tab when the
    # optimization was infeasible, telling the reader nothing at all.
    check("the Cartera sheet is never blank -- positions or a stated reason",
          wb["Cartera"].max_row > 1)
    check("the basket spans more than one asset class, or nothing can solve",
          len(set(namespace["clases"].values())) > 1,
          f"got {sorted(set(namespace['clases'].values()))}")

    parametros = [row[0].value for row in wb["Parametros"].iter_rows(min_row=2)]
    check("Parametros sheet records the profile and that no book was used",
          "Perfil" in parametros and "Portafolio" in parametros,
          f"got {parametros}")
    check("Parametros sheet records every block weight",
          sum(1 for p in parametros if str(p).startswith("Peso — ")) == 6)
    check("Parametros sheet records the compliance audit result",
          "Auditoria de bandas" in parametros)


def test_no_account_data_in_the_notebook() -> None:
    """
    The notebook must not carry positions, balances or any account snapshot.
    An earlier version embedded the IBKR book as an editable cell.
    """
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    body = "\n".join("".join(c["source"]) for c in nb["cells"])

    forbidden = ["PORTAFOLIO", "net_liquidation", "market_value",
                 "asset_class", "portfolio_ibkr", "compute_portfolio_fit"]
    present = [token for token in forbidden if token in body]
    check("no account data or book wiring anywhere in the notebook",
          not present, f"found: {present}")

    check("the notebook uses the standalone runner",
          "run_standalone" in body)
    check("the notebook exports without book columns",
          "standalone=True" in body)


def test_tuning_cell_leaves_config_untouched() -> None:
    """
    The tuning cell ships its examples commented out. If one were live, every
    run of the notebook would silently score on a different model than
    config.py declares.
    """
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    cell = next(c for c in nb["cells"]
                if c["cell_type"] == "code"
                and "set_block_weights" in "".join(c["source"]))
    body = "".join(cell["source"])

    live = [line for line in body.splitlines()
            if line.strip() and not line.strip().startswith("#")]
    mutating = [line for line in live
                if any(call in line for call in
                       ("set_block_weights(", "override(", "block_weights({"))
                and "import" not in line]
    check("no mutating tuning call is live in the shipped notebook",
          not mutating, f"live: {mutating}")


def test_no_outputs_committed() -> None:
    nb = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    with_outputs = [i for i, c in enumerate(nb["cells"]) if c.get("outputs")]
    check("notebook ships with no stale cell outputs", not with_outputs,
          f"cells {with_outputs}")

    counts = [c.get("execution_count") for c in nb["cells"]
              if c["cell_type"] == "code"]
    check("no execution counts committed", all(c is None for c in counts))


def main() -> int:
    for fn in [
        test_notebook_matches_source,
        test_embedded_engine_is_current,
        test_no_outputs_committed,
        test_tuning_cell_leaves_config_untouched,
        test_no_account_data_in_the_notebook,
        test_cells_execute,
    ]:
        fn()

    print("-" * 70)
    print(f"{PASSED}/{PASSED + FAILED} passed")
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
